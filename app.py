from flask import Flask, render_template, request, jsonify, Response
import requests, os, re
from datetime import datetime, timedelta
from io import BytesIO

app = Flask(__name__)
VT_API_KEY = os.environ.get("VT_API_KEY", "f2c83df5f4ce4ee2126f44d0082509efb1ff87aee930dffaf0772f08775d6458")
request_times = []

def detect_ioc_type(value):
    value = value.strip().lower()
    if re.match(r"^[a-f0-9]{64}$", value): return "sha256"
    elif re.match(r"^[a-f0-9]{32}$", value): return "md5"
    elif re.match(r"^[a-f0-9]{40}$", value): return "sha1"
    elif re.match(r"^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$", value): return "ip"
    elif "." in value: return "domain"
    return "unknown"

def check_rate_limit():
    global request_times
    now = datetime.now()
    cutoff = now - timedelta(minutes=1)
    request_times = [t for t in request_times if t > cutoff]
    if len(request_times) >= 4:
        return (min(request_times) + timedelta(minutes=1) - now).total_seconds()
    return 0

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/scan_one", methods=["POST"])
def scan_one():
    wait = check_rate_limit()
    if wait > 0:
        return jsonify({"status": "wait", "seconds": int(wait) + 1})
    data = request.get_json()
    value = data.get("ioc", "").strip().lower()
    ioc_type = detect_ioc_type(value)
    request_times.append(datetime.now())
    headers = {"x-apikey": VT_API_KEY}
    try:
        if ioc_type in ["md5", "sha1", "sha256"]:
            url = f"https://www.virustotal.com/api/v3/files/{value}"
        elif ioc_type == "ip":
            url = f"https://www.virustotal.com/api/v3/ip_addresses/{value}"
        elif ioc_type == "domain":
            url = f"https://www.virustotal.com/api/v3/domains/{value}"
        else:
            return jsonify({"status": "error", "error": "Tipo no soportado"})
        resp = requests.get(url, headers=headers, timeout=30)
        if resp.status_code == 404:
            return jsonify({"status": "not_found", "ioc": value, "ioc_type": ioc_type})
        if resp.status_code != 200:
            return jsonify({"status": "error", "error": f"VT error {resp.status_code}"})
        attr = resp.json().get("data", {}).get("attributes", {})
        stats = attr.get("last_analysis_stats", {})
        malicious = stats.get("malicious", 0)
        suspicious = stats.get("suspicious", 0)
        total = sum(stats.values())
        threat_label = ""
        tc = attr.get("popular_threat_classification", {})
        if tc:
            threat_label = tc.get("suggested_threat_label", "")
        return jsonify({
            "status": "scanned", "ioc": value, "ioc_type": ioc_type,
            "malicious": malicious, "suspicious": suspicious, "total": total,
            "score": f"{malicious}/{total}", "threat_label": threat_label,
            "file_name": attr.get("meaningful_name", ""),
            "file_type": attr.get("type_description", ""),
            "country": attr.get("country", ""),
            "as_owner": attr.get("as_owner", ""),
        })
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)})

@app.route("/export/xlsx", methods=["POST"])
def export_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    data = request.get_json()
    iocs = data.get("iocs", [])

    wb = Workbook()
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="182B47", end_color="182B47", fill_type="solid")
    red   = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    yellow= PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
    green = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    lred  = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    lyellow=PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
    lgreen= PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    white = Font(bold=True, color="FFFFFF")
    link  = Font(color="2980B9", underline="single")
    border= Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    total  = len(iocs)
    mal    = sum(1 for i in iocs if i.get("status")=="scanned" and i.get("malicious",0)>10)
    sus    = sum(1 for i in iocs if i.get("status")=="scanned" and 0<i.get("malicious",0)<=10)
    clean  = sum(1 for i in iocs if i.get("status")=="scanned" and i.get("malicious",0)==0)
    nf     = sum(1 for i in iocs if i.get("status")=="not_found")
    sha256c= sum(1 for i in iocs if i.get("ioc_type")=="sha256")
    md5c   = sum(1 for i in iocs if i.get("ioc_type")=="md5")
    sha1c  = sum(1 for i in iocs if i.get("ioc_type")=="sha1")
    ipc    = sum(1 for i in iocs if i.get("ioc_type")=="ip")
    domc   = sum(1 for i in iocs if i.get("ioc_type")=="domain")
    scanned= [i for i in iocs if i.get("status")=="scanned"]
    sc0    = sum(1 for i in scanned if i.get("malicious",0)==0)
    sc1    = sum(1 for i in scanned if 0<i.get("malicious",0)<=10)
    sc2    = sum(1 for i in scanned if 10<i.get("malicious",0)<=30)
    sc3    = sum(1 for i in scanned if i.get("malicious",0)>30)
    pc     = {}
    for i in iocs:
        c = i.get("country","")
        if c: pc[c] = pc.get(c,0)+1
    ps = sorted(pc.items(), key=lambda x: x[1], reverse=True)[:8]

    def pct(n): return f"{n/total*100:.1f}%" if total else "0%"

    # Hoja Resumen
    ws = wb.active
    ws.title = "Resumen Ejecutivo"
    ws.merge_cells("A1:C1")
    ws["A1"] = "REPORTE IOCs - VIRUSTOTAL - CIR Banxico"
    ws["A1"].font = Font(bold=True, size=14, color="182B47")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A3"] = "Fecha:"; ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws["A4"] = "Generado por:"; ws["B4"] = "VT-Automatizado v2"

    ws["A6"] = "ESTADISTICAS GENERALES"; ws["A6"].font = Font(bold=True)
    for c,h in enumerate(["Categoria","Cantidad","Porcentaje"],1):
        cell=ws.cell(row=7,column=c,value=h); cell.font=hf; cell.fill=hfill; cell.border=border

    rows=[("Total IOCs",total,"100%",None),("Maliciosos (>10)",mal,pct(mal),lred),
          ("Sospechosos (1-10)",sus,pct(sus),lyellow),("Limpios (0)",clean,pct(clean),lgreen),
          ("Sin resultado",nf,pct(nf),None)]
    for r,(cat,cant,p,fill) in enumerate(rows,8):
        ws.cell(row=r,column=1,value=cat).border=border
        ws.cell(row=r,column=2,value=cant).border=border
        ws.cell(row=r,column=3,value=p).border=border
        if fill: ws.cell(row=r,column=1).fill=fill

    ws["A14"]="POR TIPO DE IOC"; ws["A14"].font=Font(bold=True)
    for c,h in enumerate(["Tipo","Cantidad"],1):
        cell=ws.cell(row=15,column=c,value=h); cell.font=hf; cell.fill=hfill; cell.border=border
    for r,(t,n) in enumerate([("SHA256",sha256c),("MD5",md5c),("SHA1",sha1c),("IP",ipc),("Dominio",domc)],16):
        ws.cell(row=r,column=1,value=t).border=border
        ws.cell(row=r,column=2,value=n).border=border

    ws["A23"]="DISTRIBUCION DE SCORE"; ws["A23"].font=Font(bold=True)
    for c,h in enumerate(["Rango","Cantidad"],1):
        cell=ws.cell(row=24,column=c,value=h); cell.font=hf; cell.fill=hfill; cell.border=border
    for r,(t,n) in enumerate([("Limpio (0)",sc0),("Bajo (1-10)",sc1),("Medio (11-30)",sc2),("Alto (30+)",sc3)],25):
        ws.cell(row=r,column=1,value=t).border=border
        ws.cell(row=r,column=2,value=n).border=border

    ws["A30"]="TOP PAISES (IPs)"; ws["A30"].font=Font(bold=True)
    for c,h in enumerate(["Pais","Cantidad"],1):
        cell=ws.cell(row=31,column=c,value=h); cell.font=hf; cell.fill=hfill; cell.border=border
    for r,(country,n) in enumerate(ps,32):
        ws.cell(row=r,column=1,value=country).border=border
        ws.cell(row=r,column=2,value=n).border=border

    ws.column_dimensions["A"].width=25
    ws.column_dimensions["B"].width=12
    ws.column_dimensions["C"].width=12

    # Hoja IOCs Detallados
    ws2 = wb.create_sheet("IOCs Detallados")
    headers=["Score","IOC","Tipo","Threat Label","Archivo","Pais/AS","Link VT"]
    for c,h in enumerate(headers,1):
        cell=ws2.cell(row=1,column=c,value=h); cell.font=hf; cell.fill=hfill; cell.border=border
    ws2.freeze_panes="A2"

    sorted_iocs=sorted(iocs, key=lambda x: x.get("malicious",0), reverse=True)
    for r,ioc in enumerate(sorted_iocs,2):
        m=ioc.get("malicious",0)
        score=ioc.get("score","") or ioc.get("status","")
        itype=ioc.get("ioc_type","")
        val=ioc.get("value","")
        if itype in ["md5","sha1","sha256"]: vt_url=f"https://www.virustotal.com/gui/file/{val}"
        elif itype=="ip": vt_url=f"https://www.virustotal.com/gui/ip-address/{val}"
        elif itype=="domain": vt_url=f"https://www.virustotal.com/gui/domain/{val}"
        else: vt_url=""

        c1=ws2.cell(row=r,column=1,value=score); c1.border=border
        ws2.cell(row=r,column=2,value=val).border=border
        ws2.cell(row=r,column=3,value=itype.upper()).border=border
        ws2.cell(row=r,column=4,value=ioc.get("threat_label","")).border=border
        fn=ioc.get("file_name",""); ft=ioc.get("file_type","")
        ws2.cell(row=r,column=5,value=" - ".join(filter(None,[fn,ft]))).border=border
        co=ioc.get("country",""); ao=ioc.get("as_owner","")
        ws2.cell(row=r,column=6,value=" ".join(filter(None,[co,ao]))).border=border
        cu=ws2.cell(row=r,column=7,value=vt_url)
        if vt_url: cu.hyperlink=vt_url; cu.font=link
        cu.border=border

        if ioc.get("status")=="scanned":
            if m>10:   c1.fill=red;    c1.font=white
            elif m>0:  c1.fill=yellow
            else:      c1.fill=green;  c1.font=white
        
        # Color fila completa suave
        row_fill=None
        if ioc.get("status")=="scanned":
            if m>10: row_fill=lred
            elif m>0: row_fill=lyellow
            else: row_fill=lgreen
        if row_fill:
            for col in range(2,7):
                ws2.cell(row=r,column=col).fill=row_fill

    ws2.column_dimensions["A"].width=10
    ws2.column_dimensions["B"].width=65
    ws2.column_dimensions["C"].width=10
    ws2.column_dimensions["D"].width=28
    ws2.column_dimensions["E"].width=30
    ws2.column_dimensions["F"].width=25
    ws2.column_dimensions["G"].width=55

    output=BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":f"attachment; filename=IOCs_VT_{datetime.now().strftime(\'%Y%m%d_%H%M\')}.xlsx"})

if __name__ == "__main__":
    app.run(debug=True)
